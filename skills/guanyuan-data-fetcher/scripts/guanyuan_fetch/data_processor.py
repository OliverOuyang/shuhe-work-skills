"""Data Processor Module

Processes MCP API responses and converts them to CSV files or Excel files.
"""

import csv
from typing import Dict, List, Any
from pathlib import Path
import json

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ProcessingError(Exception):
    """Exception raised during data processing"""
    pass


def parse_mcp_response(response_data: Dict) -> tuple[List[str], List[List[Any]]]:
    """Parse MCP get_card_data response into headers and rows

    Args:
        response_data: MCP response dictionary with 'data' field containing
                      'headers' and 'values' arrays

    Returns:
        Tuple of (headers, rows) where:
        - headers: List of column names
        - rows: List of row data (each row is a list of values)

    Raises:
        ProcessingError: If response structure is invalid
    """
    try:
        # Handle nested response structure: response.data.data
        if "data" in response_data and "data" in response_data["data"]:
            data = response_data["data"]["data"]
        else:
            data = response_data.get("data", {})

        # Extract headers
        headers_obj = data.get("headers", [])
        if not headers_obj:
            raise ProcessingError("Response missing 'headers'")

        # Headers can be list of strings or list of dicts with 'name' key
        if isinstance(headers_obj[0], dict):
            headers = [h.get("name", f"column_{i}") for i, h in enumerate(headers_obj)]
        else:
            headers = headers_obj

        # Extract values
        values = data.get("values", [])
        if not isinstance(values, list):
            raise ProcessingError("'values' must be a list")

        return headers, values

    except (KeyError, IndexError, AttributeError) as e:
        raise ProcessingError(f"Invalid response structure: {e}")


def write_csv_utf8_bom(
    output_path: Path,
    headers: List[str],
    rows: List[List[Any]],
) -> None:
    """Write CSV file with UTF-8 BOM encoding

    Args:
        output_path: Path to output CSV file
        headers: List of column names
        rows: List of row data

    Raises:
        ProcessingError: If file write fails
    """
    try:
        output_path.parent.mkdir(parents=True, exist_ok=True)

        with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(rows)

    except IOError as e:
        raise ProcessingError(f"Failed to write CSV: {e}")


def process_and_save(
    response_data: Dict,
    output_filename: str,
    output_dir: Path = Path("Data/guanyuan"),
) -> Dict[str, Any]:
    """Process MCP response and save to CSV

    Args:
        response_data: MCP get_card_data response
        output_filename: Name of output CSV file
        output_dir: Directory to save CSV (default: Data/guanyuan)

    Returns:
        Dictionary with processing summary:
        {
            "output_path": str,
            "row_count": int,
            "column_count": int,
            "headers": List[str]
        }

    Raises:
        ProcessingError: If processing or saving fails
    """
    # Parse response
    headers, rows = parse_mcp_response(response_data)

    # Validate data
    if not rows:
        raise ProcessingError("No data rows returned")

    # Build output path
    output_path = output_dir / output_filename

    # Write CSV
    write_csv_utf8_bom(output_path, headers, rows)

    # Return summary
    return {
        "output_path": str(output_path),
        "row_count": len(rows),
        "column_count": len(headers),
        "headers": headers,
    }


def validate_csv_output(csv_path: Path) -> bool:
    """Validate that a CSV file was created successfully

    Args:
        csv_path: Path to CSV file

    Returns:
        True if file exists and is readable

    Raises:
        ProcessingError: If validation fails
    """
    if not csv_path.exists():
        raise ProcessingError(f"CSV file not found: {csv_path}")

    # Try to read first line to verify it's valid CSV
    try:
        with open(csv_path, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            headers = next(reader, None)
            if not headers:
                raise ProcessingError(f"CSV file is empty: {csv_path}")
    except IOError as e:
        raise ProcessingError(f"Cannot read CSV file: {e}")

    return True


def write_excel_multi_sheet(
    output_path: Path,
    cards_data: List[Dict[str, Any]],
) -> None:
    """Write multiple cards to Excel file with separate sheets

    Args:
        output_path: Path to output Excel file
        cards_data: List of card data dictionaries with structure:
                   [{"sheet_name": str, "headers": List, "rows": List}]

    Raises:
        ProcessingError: If openpyxl is not installed or file write fails
    """
    if not HAS_OPENPYXL:
        raise ProcessingError(
            "openpyxl is not installed. "
            "Install it with: pip install openpyxl"
        )

    if not cards_data:
        raise ProcessingError("No card data provided")

    try:
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Create workbook
        wb = Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        for card_data in cards_data:
            sheet_name = card_data.get("sheet_name", "Sheet1")
            headers = card_data.get("headers", [])
            rows = card_data.get("rows", [])

            # Truncate sheet name to 31 characters (Excel limit)
            sheet_name = sheet_name[:31]

            # Create sheet
            ws = wb.create_sheet(title=sheet_name)

            # Write headers
            if headers:
                ws.append(headers)

            # Write rows
            for row in rows:
                ws.append(row)

            # Auto-adjust column widths (optional, basic implementation)
            for idx, col in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
                column_letter = get_column_letter(idx)
                # Set a reasonable default width
                ws.column_dimensions[column_letter].width = 15

        # Save workbook
        wb.save(output_path)

    except IOError as e:
        raise ProcessingError(f"Failed to write Excel file: {e}")
    except Exception as e:
        raise ProcessingError(f"Unexpected error writing Excel: {e}")


def process_and_save_excel(
    cards_responses: List[Dict],
    cards_config: List,  # List[CardConfig]
    output_filename: str,
    output_dir: Path = Path("Data/guanyuan"),
) -> Dict[str, Any]:
    """Process multiple card responses and save to single Excel file

    Args:
        cards_responses: List of MCP get_card_data responses
        cards_config: List of CardConfig objects corresponding to responses
        output_filename: Name of output Excel file (should end with .xlsx)
        output_dir: Directory to save Excel (default: Data/guanyuan)

    Returns:
        Dictionary with processing summary:
        {
            "output_path": str,
            "total_rows": int,
            "sheet_count": int,
            "sheets": List[Dict] # per-sheet stats
        }

    Raises:
        ProcessingError: If processing or saving fails
    """
    if len(cards_responses) != len(cards_config):
        raise ProcessingError(
            f"Mismatched lengths: {len(cards_responses)} responses "
            f"vs {len(cards_config)} configs"
        )

    if not output_filename.endswith(".xlsx"):
        output_filename = output_filename.replace(".csv", ".xlsx")
        if not output_filename.endswith(".xlsx"):
            output_filename += ".xlsx"

    # Process all cards
    cards_data = []
    sheets_info = []
    total_rows = 0

    for response, config in zip(cards_responses, cards_config):
        try:
            # Parse response
            headers, rows = parse_mcp_response(response)

            # Validate data
            if not rows:
                # Allow empty cards, but log it
                rows = []

            # Use card_name or tab_name as sheet name
            sheet_name = getattr(config, "card_name", "Sheet")

            cards_data.append({
                "sheet_name": sheet_name,
                "headers": headers,
                "rows": rows,
            })

            # Track stats
            row_count = len(rows)
            total_rows += row_count
            sheets_info.append({
                "sheet_name": sheet_name[:31],
                "row_count": row_count,
                "column_count": len(headers),
            })

        except ProcessingError as e:
            # Re-raise with card context
            raise ProcessingError(
                f"Failed to process card '{getattr(config, 'card_name', 'unknown')}': {e}"
            )

    # Build output path
    output_path = output_dir / output_filename

    # Write Excel
    write_excel_multi_sheet(output_path, cards_data)

    # Return summary
    return {
        "output_path": str(output_path),
        "total_rows": total_rows,
        "sheet_count": len(cards_data),
        "sheets": sheets_info,
    }

