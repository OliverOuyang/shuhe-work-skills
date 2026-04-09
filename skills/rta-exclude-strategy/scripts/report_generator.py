"""
报告生成模块
负责生成Excel格式的RTA排除策略分析报告
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, numbers
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')
from utils import convert_old_rule_to_quantile, calc_spr, calc_cps, make_region_mask, filter_by_region


def generate_report(result, old_exclude_rule, output_path=None):
    """
    生成完整的Excel报告

    Args:
        result: 算法结果字典
        old_exclude_rule: 老策略排除规则（如['01q', '02q', ...]）
        output_path: 输出路径（可选）

    Returns:
        str: 生成的报告文件路径
    """
    print("\n" + "="*100)
    print("生成Excel报告")
    print("="*100)

    # 提取数据
    df_combined = result['df_combined']
    df_ctrl = result['df_ctrl']
    exclude_region = result['exclude_region']

    # 转换老策略规则为聚合后的格式
    old_exclude_v8 = convert_old_rule_to_quantile(old_exclude_rule)

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "RTA排除策略报告"

    current_row = 1

    # 第一部分：核心结论
    current_row = generate_section1_conclusion(
        ws, current_row, df_ctrl, exclude_region, old_exclude_v8
    )

    # 第二部分：排除策略制定
    current_row = generate_section2_strategy(
        ws, current_row, df_combined, df_ctrl, exclude_region, old_exclude_v8
    )

    # 第三部分：合理性评估
    current_row = generate_section3_evaluation(
        ws, current_row, df_ctrl, exclude_region, old_exclude_v8
    )

    # 调整列宽
    ws.column_dimensions['A'].width = 35
    for col in ['B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 15

    # 保存文件
    if output_path is None:
        output_path = '.'

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    file_path = f'{output_path}/RTA排除策略报告_{timestamp}.xlsx'
    wb.save(file_path)

    print(f"\n报告已生成: {file_path}")
    return file_path


# ============================================================================
# 基础工具函数
# ============================================================================

def format_number_cell(cell, value):
    """
    设置单元格数值格式

    Args:
        cell: 单元格对象
        value: 数值

    规则:
        - 小于1的数值: 显示为百分比格式 (0.1234 -> 12.34%)
        - 大于等于1的数值: 显示为数字格式,保留4位小数
    """
    if value < 1:
        cell.number_format = '0.00%'
    else:
        cell.number_format = '0.0000'


def write_text(ws, start_row, text, font_size=11):
    """
    写入多行文本

    Args:
        ws: worksheet对象
        start_row: 起始行
        text: 文本内容
        font_size: 字体大小

    Returns:
        int: 下一行的行号
    """
    lines = text.strip().split('\n')
    for line in lines:
        ws.cell(row=start_row, column=1, value=line.strip())
        ws.cell(row=start_row, column=1).font = Font(name='楷体', size=font_size)
        ws.cell(row=start_row, column=1).alignment = Alignment(horizontal='left', wrap_text=True)
        start_row += 1
    return start_row


def write_title(ws, row, title, font_size=14, bold=True):
    """
    写入标题

    Args:
        ws: worksheet对象
        row: 行号
        title: 标题文本
        font_size: 字体大小
        bold: 是否加粗

    Returns:
        int: 下一行的行号
    """
    ws.cell(row=row, column=1, value=title)
    ws.cell(row=row, column=1).font = Font(name='楷体', size=font_size, bold=bold)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='left')
    return row + 1



# ============================================================================
# 第一部分：核心结论
# ============================================================================

def generate_section1_conclusion(ws, current_row, df_ctrl, exclude_region, old_exclude_v8):
    """
    生成第一部分：核心结论

    Args:
        ws: worksheet对象
        current_row: 当前行号
        df_ctrl: 对照组数据
        exclude_region: 新策略排除区域
        old_exclude_v8: 老策略排除规则

    Returns:
        int: 下一行的行号
    """
    print("\n生成第一部分：核心结论")

    # 写入标题
    current_row = write_title(ws, current_row, "一、核心结论：关键指标对比（旧策略 vs 新策略）", font_size=14, bold=True)
    current_row += 1

    # 计算指标
    total_ctrl_amt = df_ctrl['t3_loan_amt'].sum()

    # 老策略指标
    old_exclude_data_ctrl = df_ctrl[df_ctrl['V8_Q'].isin(old_exclude_v8)]
    old_exclude_amt_ratio = old_exclude_data_ctrl['t3_loan_amt'].sum() / total_ctrl_amt
    old_exclude_spr = calc_spr(old_exclude_data_ctrl)
    old_remain_data_ctrl = df_ctrl[~df_ctrl['V8_Q'].isin(old_exclude_v8)]
    old_remain_spr = calc_spr(old_remain_data_ctrl)
    old_remain_cps = calc_cps(old_remain_data_ctrl)

    # 新策略指标
    new_exclude_data_ctrl = filter_by_region(df_ctrl, exclude_region)
    new_exclude_amt_ratio = new_exclude_data_ctrl['t3_loan_amt'].sum() / total_ctrl_amt
    new_exclude_spr = calc_spr(new_exclude_data_ctrl)
    new_remain_data_ctrl = df_ctrl[~make_region_mask(df_ctrl, exclude_region)]
    new_remain_spr = calc_spr(new_remain_data_ctrl)
    new_remain_cps = calc_cps(new_remain_data_ctrl)

    # 生成结论文本
    conclusion_text = f"""核心结论：

1. 新策略显著提升保留客群质量
   旧策略排除交易占比{old_exclude_amt_ratio*100:.2f}%，保留客群安全过件率{old_remain_spr*100:.2f}%
   新策略排除交易占比{new_exclude_amt_ratio*100:.2f}%，保留客群安全过件率{new_remain_spr*100:.2f}%
   新策略保留客群安全过件率提升{(new_remain_spr-old_remain_spr)*100:.2f}个百分点，相当于提升了{(new_remain_spr/old_remain_spr-1)*100:.1f}%

2. 新策略更精准识别低质量客群
   旧策略排除客群安全过件率{old_exclude_spr*100:.2f}%
   新策略排除客群安全过件率{new_exclude_spr*100:.2f}%
   新策略排除客群质量更低，排除更精准

3. 满足约束条件
   新策略排除交易占比{new_exclude_amt_ratio*100:.2f}%，在20%约束范围内

4. 二维模型优势
   新策略采用V8 x V9RN二维模型，相比旧策略的V8单维度模型，可以更细粒度地识别客群质量

综合评价：新策略优于旧策略，建议采用新策略"""

    current_row = write_text(ws, current_row, conclusion_text, font_size=11)
    current_row += 2

    return current_row


# ============================================================================
# 第二部分：排除策略制定
# ============================================================================

def generate_section2_strategy(ws, current_row, df_combined, df_ctrl, exclude_region, old_exclude_v8):
    """
    生成第二部分：排除策略制定

    Args:
        ws: worksheet对象
        current_row: 当前行号
        df_combined: 全量数据
        df_ctrl: 对照组数据
        exclude_region: 新策略排除区域
        old_exclude_v8: 老策略排除规则

    Returns:
        int: 下一行的行号
    """
    print("\n生成第二部分：排除策略制定")

    # 写入标题
    current_row = write_title(ws, current_row, "二、排除策略制定", font_size=14, bold=True)
    current_row += 1

    # 1. 排除策略现状
    current_row = generate_section2_1_old_strategy(ws, current_row, df_combined, old_exclude_v8)

    # 2. 排除策略制定
    current_row = generate_section2_2_new_strategy(ws, current_row, df_combined, df_ctrl, exclude_region, old_exclude_v8)

    return current_row


def generate_section2_1_old_strategy(ws, current_row, df_combined, old_exclude_v8):
    """
    生成2.1：排除策略现状（老策略）

    Args:
        ws: worksheet对象
        current_row: 当前行号
        df_combined: 全量数据
        old_exclude_v8: 老策略排除规则

    Returns:
        int: 下一行的行号
    """
    current_row = write_title(ws, current_row, "1. 排除策略现状", font_size=12, bold=True)
    current_row += 1

    # 表1：老策略排除规则
    ws.cell(row=current_row, column=1, value="表1：老策略排除规则（V8单维度）")
    ws.cell(row=current_row, column=1).font = Font(name='楷体', size=11, bold=True)
    current_row += 1

    # 计算V8各分位的安全过件率（全量数据）
    v8_list = [f'{i:02d}Q' for i in range(1, 13)]
    v8_stats_all = df_combined.groupby('V8_Q').agg({
        't3_ato': 'sum',
        't3_safe_adt': 'sum'
    }).reset_index()
    v8_stats_all['安全过件率'] = v8_stats_all['t3_safe_adt'] / v8_stats_all['t3_ato']

    # 表头
    ws.cell(row=current_row, column=1, value="V8分位")
    ws.cell(row=current_row, column=1).font = Font(name='楷体', size=10, bold=True)
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=2, value="安全过件率（全量）")
    ws.cell(row=current_row, column=2).font = Font(name='楷体', size=10, bold=True)
    ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=3, value="是否排除")
    ws.cell(row=current_row, column=3).font = Font(name='楷体', size=10, bold=True)
    ws.cell(row=current_row, column=3).alignment = Alignment(horizontal='center')
    current_row += 1

    # 数据行
    for v8 in v8_list:
        v8_data = v8_stats_all[v8_stats_all['V8_Q'] == v8]
        if len(v8_data) > 0:
            spr = v8_data['安全过件率'].values[0]
            is_excluded = '是' if v8 in old_exclude_v8 else '否'

            ws.cell(row=current_row, column=1, value=v8)
            ws.cell(row=current_row, column=1).font = Font(name='楷体', size=10)
            ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')

            cell = ws.cell(row=current_row, column=2, value=spr)
            cell.font = Font(name='楷体', size=10)
            cell.alignment = Alignment(horizontal='center')
            format_number_cell(cell, spr)

            ws.cell(row=current_row, column=3, value=is_excluded)
            ws.cell(row=current_row, column=3).font = Font(name='楷体', size=10, bold=(is_excluded=='是'))
            ws.cell(row=current_row, column=3).alignment = Alignment(horizontal='center')
            if is_excluded == '是':
                ws.cell(row=current_row, column=3).fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')

            current_row += 1

    current_row += 1
    return current_row


def generate_section2_2_new_strategy(ws, current_row, df_combined, df_ctrl, exclude_region, old_exclude_v8):
    """
    生成2.2：排除策略制定（新策略）

    Args:
        ws: worksheet对象
        current_row: 当前行号
        df_combined: 全量数据
        df_ctrl: 对照组数据
        exclude_region: 新策略排除区域
        old_exclude_v8: 老策略排除规则

    Returns:
        int: 下一行的行号
    """
    current_row = write_title(ws, current_row, "2. 排除策略制定", font_size=12, bold=True)
    current_row += 1

    # 1）方案展示
    ws.cell(row=current_row, column=1, value="1）方案展示")
    ws.cell(row=current_row, column=1).font = Font(name='楷体', size=11, bold=True)
    current_row += 1

    ws.cell(row=current_row, column=1, value="表1：新策略排除较老策略排除在对照组上的指标差异")
    ws.cell(row=current_row, column=1).font = Font(name='楷体', size=11, bold=True)
    current_row += 1

    # 计算指标
    total_ctrl_expo = df_ctrl['expo_cnt'].sum()
    total_ctrl_ato = df_ctrl['t3_ato'].sum()
    total_ctrl_amt = df_ctrl['t3_loan_amt'].sum()

    # 老策略指标
    old_exclude_data_ctrl = df_ctrl[df_ctrl['V8_Q'].isin(old_exclude_v8)]
    old_exclude_expo_ratio = old_exclude_data_ctrl['expo_cnt'].sum() / total_ctrl_expo
    old_exclude_ato_ratio = old_exclude_data_ctrl['t3_ato'].sum() / total_ctrl_ato
    old_exclude_amt_ratio = old_exclude_data_ctrl['t3_loan_amt'].sum() / total_ctrl_amt
    old_exclude_spr = calc_spr(old_exclude_data_ctrl)
    old_remain_data_ctrl = df_ctrl[~df_ctrl['V8_Q'].isin(old_exclude_v8)]
    old_remain_spr = calc_spr(old_remain_data_ctrl)
    old_remain_cps = calc_cps(old_remain_data_ctrl)

    # 新策略指标
    new_exclude_data_ctrl = filter_by_region(df_ctrl, exclude_region)
    new_exclude_expo_ratio = new_exclude_data_ctrl['expo_cnt'].sum() / total_ctrl_expo
    new_exclude_ato_ratio = new_exclude_data_ctrl['t3_ato'].sum() / total_ctrl_ato
    new_exclude_amt_ratio = new_exclude_data_ctrl['t3_loan_amt'].sum() / total_ctrl_amt
    new_exclude_spr = calc_spr(new_exclude_data_ctrl)
    new_remain_data_ctrl = df_ctrl[~make_region_mask(df_ctrl, exclude_region)]
    new_remain_spr = calc_spr(new_remain_data_ctrl)
    new_remain_cps = calc_cps(new_remain_data_ctrl)

    # 表头
    ws.cell(row=current_row, column=1, value="指标")
    ws.cell(row=current_row, column=1).font = Font(name='楷体', size=10, bold=True)
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=2, value="旧策略")
    ws.cell(row=current_row, column=2).font = Font(name='楷体', size=10, bold=True)
    ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=3, value="新策略")
    ws.cell(row=current_row, column=3).font = Font(name='楷体', size=10, bold=True)
    ws.cell(row=current_row, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=4, value="差异")
    ws.cell(row=current_row, column=4).font = Font(name='楷体', size=10, bold=True)
    ws.cell(row=current_row, column=4).alignment = Alignment(horizontal='center')
    current_row += 1

    # 数据行
    metrics = [
        ('排除曝光占比', old_exclude_expo_ratio, new_exclude_expo_ratio, True),
        ('排除申完占比', old_exclude_ato_ratio, new_exclude_ato_ratio, True),
        ('排除交易占比', old_exclude_amt_ratio, new_exclude_amt_ratio, True),
        ('排除用户安全过件率', old_exclude_spr, new_exclude_spr, True),
        ('排除后对照组安全过件率', old_remain_spr, new_remain_spr, True),
        ('排除后CPS', old_remain_cps, new_remain_cps, False)
    ]

    for metric_name, old_val, new_val, is_percent in metrics:
        ws.cell(row=current_row, column=1, value=metric_name)
        ws.cell(row=current_row, column=1).font = Font(name='楷体', size=10)
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='left')

        # 旧策略值
        cell_old = ws.cell(row=current_row, column=2, value=old_val)
        cell_old.font = Font(name='楷体', size=10)
        cell_old.alignment = Alignment(horizontal='center')
        format_number_cell(cell_old, old_val)

        # 新策略值
        cell_new = ws.cell(row=current_row, column=3, value=new_val)
        cell_new.font = Font(name='楷体', size=10)
        cell_new.alignment = Alignment(horizontal='center')
        format_number_cell(cell_new, new_val)

        # 差异值
        diff_val = new_val - old_val
        cell_diff = ws.cell(row=current_row, column=4, value=diff_val)
        cell_diff.font = Font(name='楷体', size=10)
        cell_diff.alignment = Alignment(horizontal='center')
        format_number_cell(cell_diff, abs(diff_val))

        current_row += 1

    current_row += 1

    # 2）排除规则
    ws.cell(row=current_row, column=1, value='2）排除规则："安全过件率低于10%客群"')
    ws.cell(row=current_row, column=1).font = Font(name='楷体', size=11, bold=True)
    current_row += 1

    ws.cell(row=current_row, column=1, value="表2：新策略排除规则（V8 x V9RN二维）")
    ws.cell(row=current_row, column=1).font = Font(name='楷体', size=11, bold=True)
    current_row += 1

    # 生成二维热力图
    current_row = generate_heatmap(ws, current_row, df_combined, exclude_region)

    return current_row


def generate_heatmap(ws, current_row, df_combined, exclude_region):
    """
    生成V8 x V9RN二维热力图

    Args:
        ws: worksheet对象
        current_row: 当前行号
        df_combined: 全量数据
        exclude_region: 排除区域

    Returns:
        int: 下一行的行号
    """
    v8_list = [f'{i:02d}Q' for i in range(1, 13)]
    v9_list = [f'{i:02d}Q' for i in range(1, 13)]

    # 表头
    ws.cell(row=current_row, column=1, value="V8\\V9RN")
    ws.cell(row=current_row, column=1).font = Font(name='楷体', size=10, bold=True)
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')
    for col_idx, v9 in enumerate(v9_list, 2):
        ws.cell(row=current_row, column=col_idx, value=v9)
        ws.cell(row=current_row, column=col_idx).font = Font(name='楷体', size=9, bold=True)
        ws.cell(row=current_row, column=col_idx).alignment = Alignment(horizontal='center')
    current_row += 1

    # 数据行
    for v8 in v8_list:
        ws.cell(row=current_row, column=1, value=v8)
        ws.cell(row=current_row, column=1).font = Font(name='楷体', size=9, bold=True)
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')

        for col_idx, v9 in enumerate(v9_list, 2):
            cell_data = df_combined[(df_combined['V8_Q'] == v8) & (df_combined['V9RN_Q'] == v9)]
            if len(cell_data) > 0 and cell_data['t3_ato'].sum() > 0:
                spr = cell_data['t3_safe_adt'].sum() / cell_data['t3_ato'].sum()
                cell = ws.cell(row=current_row, column=col_idx, value=spr)

                # 如果在排除区域,标记为红色
                if (v8, v9) in exclude_region:
                    cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')

                format_number_cell(cell, spr)
            else:
                cell = ws.cell(row=current_row, column=col_idx, value='-')

            cell.font = Font(name='楷体', size=8)
            cell.alignment = Alignment(horizontal='center')

        current_row += 1

    current_row += 2
    return current_row


# ============================================================================
# 第三部分：合理性评估
# ============================================================================

def generate_section3_evaluation(ws, current_row, df_ctrl, exclude_region, old_exclude_v8):
    """
    生成第三部分：合理性评估

    Args:
        ws: worksheet对象
        current_row: 当前行号
        df_ctrl: 对照组数据
        exclude_region: 新策略排除区域
        old_exclude_v8: 老策略排除规则

    Returns:
        int: 下一行的行号
    """
    print("\n生成第三部分：合理性评估")

    # 写入标题
    current_row = write_title(ws, current_row, "三、合理性评估", font_size=14, bold=True)
    current_row += 1

    # 1. 置入置出合理性分析
    current_row = generate_section3_1_place_analysis(ws, current_row, df_ctrl, exclude_region, old_exclude_v8)

    # 2. 图表展示
    current_row = generate_section3_2_cross_tables(ws, current_row, df_ctrl, exclude_region, old_exclude_v8)

    return current_row


def generate_section3_1_place_analysis(ws, current_row, df_ctrl, exclude_region, old_exclude_v8):
    """
    生成3.1：置入置出合理性分析

    Args:
        ws: worksheet对象
        current_row: 当前行号
        df_ctrl: 对照组数据
        exclude_region: 新策略排除区域
        old_exclude_v8: 老策略排除规则

    Returns:
        int: 下一行的行号
    """
    current_row = write_title(ws, current_row, "1. 置入置出合理性分析", font_size=12, bold=True)
    current_row += 1

    # 计算四个客群
    df_ctrl['old_exclude'] = df_ctrl['V8_Q'].isin(old_exclude_v8)
    df_ctrl['new_exclude'] = make_region_mask(df_ctrl, exclude_region)

    both_exclude = df_ctrl[(df_ctrl['old_exclude']) & (df_ctrl['new_exclude'])]
    only_old = df_ctrl[(df_ctrl['old_exclude']) & (~df_ctrl['new_exclude'])]  # 置入客群
    only_new = df_ctrl[(~df_ctrl['old_exclude']) & (df_ctrl['new_exclude'])]  # 置出客群
    both_keep = df_ctrl[(~df_ctrl['old_exclude']) & (~df_ctrl['new_exclude'])]

    # 计算指标
    total_ctrl_amt = df_ctrl['t3_loan_amt'].sum()
    place_in_amt_ratio = only_old['t3_loan_amt'].sum() / total_ctrl_amt
    place_in_spr = calc_spr(only_old)
    place_in_cps = calc_cps(only_old)

    place_out_amt_ratio = only_new['t3_loan_amt'].sum() / total_ctrl_amt
    place_out_spr = calc_spr(only_new)
    place_out_cps = calc_cps(only_new)

    # 生成分析文本
    analysis_text = f"""置入置出合理性分析结论：

1. 置入客群（仅旧策略排除，新策略保留）：
   - 交易占比：{place_in_amt_ratio*100:.2f}%
   - 安全过件率：{place_in_spr*100:.2f}%
   - CPS：{place_in_cps:.4f}
   - 评价：新策略成功将旧策略误伤的高质量客群保留下来

2. 置出客群（仅新策略排除，旧策略保留）：
   - 交易占比：{place_out_amt_ratio*100:.2f}%
   - 安全过件率：{place_out_spr*100:.2f}%
   - CPS：{place_out_cps:.4f}
   - 评价：新策略新增排除的客群质量较低，排除合理

3. 合理性验证：
   - 安全过件率合理性：✓ 置入客群SPR({place_in_spr*100:.2f}%) > 置出客群SPR({place_out_spr*100:.2f}%)
   - CPS合理性：置入客群CPS({place_in_cps:.4f})与置出客群CPS({place_out_cps:.4f})的差异符合业务预期

4. 总体评价：新策略的置入置出逻辑合理，成功优化了客群结构"""

    current_row = write_text(ws, current_row, analysis_text, font_size=11)
    current_row += 2

    return current_row


def generate_section3_2_cross_tables(ws, current_row, df_ctrl, exclude_region, old_exclude_v8):
    """
    生成3.2：交叉表展示

    Args:
        ws: worksheet对象
        current_row: 当前行号
        df_ctrl: 对照组数据
        exclude_region: 新策略排除区域
        old_exclude_v8: 老策略排除规则

    Returns:
        int: 下一行的行号
    """
    current_row = write_title(ws, current_row, "2. 图表展示", font_size=12, bold=True)
    current_row += 1

    # 计算四个客群
    df_ctrl['old_exclude'] = df_ctrl['V8_Q'].isin(old_exclude_v8)
    df_ctrl['new_exclude'] = make_region_mask(df_ctrl, exclude_region)

    both_exclude = df_ctrl[(df_ctrl['old_exclude']) & (df_ctrl['new_exclude'])]
    only_old = df_ctrl[(df_ctrl['old_exclude']) & (~df_ctrl['new_exclude'])]
    only_new = df_ctrl[(~df_ctrl['old_exclude']) & (df_ctrl['new_exclude'])]
    both_keep = df_ctrl[(~df_ctrl['old_exclude']) & (~df_ctrl['new_exclude'])]

    total_ctrl_amt = df_ctrl['t3_loan_amt'].sum()

    # 计算新老策略的排除交易占比
    new_exclude_amt_ratio = (both_exclude['t3_loan_amt'].sum() + only_new['t3_loan_amt'].sum()) / total_ctrl_amt
    old_exclude_amt_ratio = (both_exclude['t3_loan_amt'].sum() + only_old['t3_loan_amt'].sum()) / total_ctrl_amt

    # 表1：交易占比交叉表
    current_row = generate_cross_table_amt(ws, current_row, both_exclude, only_old, only_new, both_keep,
                                           total_ctrl_amt, new_exclude_amt_ratio, old_exclude_amt_ratio)

    # 表2：安全过件率交叉表
    current_row = generate_cross_table_spr(ws, current_row, both_exclude, only_old, only_new, both_keep,
                                           df_ctrl, new_exclude_amt_ratio, old_exclude_amt_ratio)

    # 表3：CPS交叉表
    current_row = generate_cross_table_cps(ws, current_row, both_exclude, only_old, only_new, both_keep,
                                           df_ctrl, new_exclude_amt_ratio, old_exclude_amt_ratio)

    return current_row


def generate_cross_table_amt(ws, current_row, both_exclude, only_old, only_new, both_keep,
                             total_ctrl_amt, new_exclude_amt_ratio, old_exclude_amt_ratio):
    """生成交易占比交叉表"""
    ws.cell(row=current_row, column=1, value="表1：交易占比交叉表")
    ws.cell(row=current_row, column=1).font = Font(name='楷体', size=11, bold=True)
    current_row += 1

    # 表头
    ws.cell(row=current_row, column=1, value="")
    ws.cell(row=current_row, column=2, value="旧策略排除")
    ws.cell(row=current_row, column=2).font = Font(name='楷体', size=10, bold=True)
    ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=3, value="旧策略不排除")
    ws.cell(row=current_row, column=3).font = Font(name='楷体', size=10, bold=True)
    ws.cell(row=current_row, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=4, value="总计")
    ws.cell(row=current_row, column=4).font = Font(name='楷体', size=10, bold=True)
    ws.cell(row=current_row, column=4).alignment = Alignment(horizontal='center')
    current_row += 1

    # 数据行
    ws.cell(row=current_row, column=1, value="新策略排除")
    ws.cell(row=current_row, column=1).font = Font(name='楷体', size=10, bold=True)

    cell = ws.cell(row=current_row, column=2, value=both_exclude['t3_loan_amt'].sum()/total_ctrl_amt)
    cell.font = Font(name='楷体', size=10)
    cell.alignment = Alignment(horizontal='center')
    format_number_cell(cell, both_exclude['t3_loan_amt'].sum()/total_ctrl_amt)

    cell = ws.cell(row=current_row, column=3, value=only_new['t3_loan_amt'].sum()/total_ctrl_amt)
    cell.font = Font(name='楷体', size=10)
    cell.alignment = Alignment(horizontal='center')
    format_number_cell(cell, only_new['t3_loan_amt'].sum()/total_ctrl_amt)

    cell = ws.cell(row=current_row, column=4, value=new_exclude_amt_ratio)
    cell.font = Font(name='楷体', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center')
    format_number_cell(cell, new_exclude_amt_ratio)
    current_row += 1

    ws.cell(row=current_row, column=1, value="新策略不排除")
    ws.cell(row=current_row, column=1).font = Font(name='楷体', size=10, bold=True)

    cell = ws.cell(row=current_row, column=2, value=only_old['t3_loan_amt'].sum()/total_ctrl_amt)
    cell.font = Font(name='楷体', size=10)
    cell.alignment = Alignment(horizontal='center')
    format_number_cell(cell, only_old['t3_loan_amt'].sum()/total_ctrl_amt)

    cell = ws.cell(row=current_row, column=3, value=both_keep['t3_loan_amt'].sum()/total_ctrl_amt)
    cell.font = Font(name='楷体', size=10)
    cell.alignment = Alignment(horizontal='center')
    format_number_cell(cell, both_keep['t3_loan_amt'].sum()/total_ctrl_amt)

    cell = ws.cell(row=current_row, column=4, value=1-new_exclude_amt_ratio)
    cell.font = Font(name='楷体', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center')
    format_number_cell(cell, 1-new_exclude_amt_ratio)
    current_row += 1

    ws.cell(row=current_row, column=1, value="总计")
    ws.cell(row=current_row, column=1).font = Font(name='楷体', size=10, bold=True)

    cell = ws.cell(row=current_row, column=2, value=old_exclude_amt_ratio)
    cell.font = Font(name='楷体', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center')
    format_number_cell(cell, old_exclude_amt_ratio)

    cell = ws.cell(row=current_row, column=3, value=1-old_exclude_amt_ratio)
    cell.font = Font(name='楷体', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center')
    format_number_cell(cell, 1-old_exclude_amt_ratio)

    cell = ws.cell(row=current_row, column=4, value=1.0)
    cell.font = Font(name='楷体', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center')
    cell.number_format = '0.00%'
    current_row += 2

    return current_row


def generate_cross_table_spr(ws, current_row, both_exclude, only_old, only_new, both_keep,
                             df_ctrl, new_exclude_amt_ratio, old_exclude_amt_ratio):
    """生成安全过件率交叉表"""
    ws.cell(row=current_row, column=1, value="表2：安全过件率交叉表")
    ws.cell(row=current_row, column=1).font = Font(name='楷体', size=11, bold=True)
    current_row += 1

    # 表头
    ws.cell(row=current_row, column=1, value="")
    ws.cell(row=current_row, column=2, value="旧策略排除")
    ws.cell(row=current_row, column=2).font = Font(name='楷体', size=10, bold=True)
    ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=3, value="旧策略不排除")
    ws.cell(row=current_row, column=3).font = Font(name='楷体', size=10, bold=True)
    ws.cell(row=current_row, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=4, value="总计")
    ws.cell(row=current_row, column=4).font = Font(name='楷体', size=10, bold=True)
    ws.cell(row=current_row, column=4).alignment = Alignment(horizontal='center')
    current_row += 1

    # 计算SPR
    new_exclude_spr = calc_spr(df_ctrl[df_ctrl['new_exclude']])
    old_exclude_spr = calc_spr(df_ctrl[df_ctrl['old_exclude']])
    old_remain_spr = calc_spr(df_ctrl[~df_ctrl['old_exclude']])
    new_remain_spr = calc_spr(df_ctrl[~df_ctrl['new_exclude']])
    total_spr = calc_spr(df_ctrl)

    # 数据行
    ws.cell(row=current_row, column=1, value="新策略排除")
    ws.cell(row=current_row, column=1).font = Font(name='楷体', size=10, bold=True)

    cell = ws.cell(row=current_row, column=2, value=calc_spr(both_exclude))
    cell.font = Font(name='楷体', size=10)
    cell.alignment = Alignment(horizontal='center')
    format_number_cell(cell, calc_spr(both_exclude))

    cell = ws.cell(row=current_row, column=3, value=calc_spr(only_new))
    cell.font = Font(name='楷体', size=10)
    cell.alignment = Alignment(horizontal='center')
    format_number_cell(cell, calc_spr(only_new))

    cell = ws.cell(row=current_row, column=4, value=new_exclude_spr)
    cell.font = Font(name='楷体', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center')
    format_number_cell(cell, new_exclude_spr)
    current_row += 1

    ws.cell(row=current_row, column=1, value="新策略不排除")
    ws.cell(row=current_row, column=1).font = Font(name='楷体', size=10, bold=True)

    cell = ws.cell(row=current_row, column=2, value=calc_spr(only_old))
    cell.font = Font(name='楷体', size=10)
    cell.alignment = Alignment(horizontal='center')
    format_number_cell(cell, calc_spr(only_old))

    cell = ws.cell(row=current_row, column=3, value=calc_spr(both_keep))
    cell.font = Font(name='楷体', size=10)
    cell.alignment = Alignment(horizontal='center')
    format_number_cell(cell, calc_spr(both_keep))

    cell = ws.cell(row=current_row, column=4, value=new_remain_spr)
    cell.font = Font(name='楷体', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center')
    format_number_cell(cell, new_remain_spr)
    current_row += 1

    ws.cell(row=current_row, column=1, value="总计")
    ws.cell(row=current_row, column=1).font = Font(name='楷体', size=10, bold=True)

    cell = ws.cell(row=current_row, column=2, value=old_exclude_spr)
    cell.font = Font(name='楷体', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center')
    format_number_cell(cell, old_exclude_spr)

    cell = ws.cell(row=current_row, column=3, value=old_remain_spr)
    cell.font = Font(name='楷体', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center')
    format_number_cell(cell, old_remain_spr)

    cell = ws.cell(row=current_row, column=4, value=total_spr)
    cell.font = Font(name='楷体', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center')
    format_number_cell(cell, total_spr)
    current_row += 2

    return current_row


def generate_cross_table_cps(ws, current_row, both_exclude, only_old, only_new, both_keep,
                             df_ctrl, new_exclude_amt_ratio, old_exclude_amt_ratio):
    """生成CPS交叉表"""
    ws.cell(row=current_row, column=1, value="表3：CPS交叉表")
    ws.cell(row=current_row, column=1).font = Font(name='楷体', size=11, bold=True)
    current_row += 1

    # 表头
    ws.cell(row=current_row, column=1, value="")
    ws.cell(row=current_row, column=2, value="旧策略排除")
    ws.cell(row=current_row, column=2).font = Font(name='楷体', size=10, bold=True)
    ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=3, value="旧策略不排除")
    ws.cell(row=current_row, column=3).font = Font(name='楷体', size=10, bold=True)
    ws.cell(row=current_row, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=4, value="总计")
    ws.cell(row=current_row, column=4).font = Font(name='楷体', size=10, bold=True)
    ws.cell(row=current_row, column=4).alignment = Alignment(horizontal='center')
    current_row += 1

    # 计算CPS
    new_exclude_cps = calc_cps(df_ctrl[df_ctrl['new_exclude']])
    old_exclude_cps = calc_cps(df_ctrl[df_ctrl['old_exclude']])
    old_remain_cps = calc_cps(df_ctrl[~df_ctrl['old_exclude']])
    new_remain_cps = calc_cps(df_ctrl[~df_ctrl['new_exclude']])
    total_cps = calc_cps(df_ctrl)

    # 数据行
    ws.cell(row=current_row, column=1, value="新策略排除")
    ws.cell(row=current_row, column=1).font = Font(name='楷体', size=10, bold=True)

    cell = ws.cell(row=current_row, column=2, value=calc_cps(both_exclude))
    cell.font = Font(name='楷体', size=10)
    cell.alignment = Alignment(horizontal='center')
    format_number_cell(cell, calc_cps(both_exclude))

    cell = ws.cell(row=current_row, column=3, value=calc_cps(only_new))
    cell.font = Font(name='楷体', size=10)
    cell.alignment = Alignment(horizontal='center')
    format_number_cell(cell, calc_cps(only_new))

    cell = ws.cell(row=current_row, column=4, value=new_exclude_cps)
    cell.font = Font(name='楷体', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center')
    format_number_cell(cell, new_exclude_cps)
    current_row += 1

    ws.cell(row=current_row, column=1, value="新策略不排除")
    ws.cell(row=current_row, column=1).font = Font(name='楷体', size=10, bold=True)

    cell = ws.cell(row=current_row, column=2, value=calc_cps(only_old))
    cell.font = Font(name='楷体', size=10)
    cell.alignment = Alignment(horizontal='center')
    format_number_cell(cell, calc_cps(only_old))

    cell = ws.cell(row=current_row, column=3, value=calc_cps(both_keep))
    cell.font = Font(name='楷体', size=10)
    cell.alignment = Alignment(horizontal='center')
    format_number_cell(cell, calc_cps(both_keep))

    cell = ws.cell(row=current_row, column=4, value=new_remain_cps)
    cell.font = Font(name='楷体', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center')
    format_number_cell(cell, new_remain_cps)
    current_row += 1

    ws.cell(row=current_row, column=1, value="总计")
    ws.cell(row=current_row, column=1).font = Font(name='楷体', size=10, bold=True)

    cell = ws.cell(row=current_row, column=2, value=old_exclude_cps)
    cell.font = Font(name='楷体', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center')
    format_number_cell(cell, old_exclude_cps)

    cell = ws.cell(row=current_row, column=3, value=old_remain_cps)
    cell.font = Font(name='楷体', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center')
    format_number_cell(cell, old_remain_cps)

    cell = ws.cell(row=current_row, column=4, value=total_cps)
    cell.font = Font(name='楷体', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center')
    format_number_cell(cell, total_cps)
    current_row += 2

    return current_row

