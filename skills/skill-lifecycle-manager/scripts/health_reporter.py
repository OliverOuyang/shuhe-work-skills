#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Health Reporter

Generates comprehensive health reports for the skills library
with Excel export support.
"""

import json
import sys
from pathlib import Path
from typing import Dict, List, Optional
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from dependency_checker import DependencyChecker
    from stats_tracker import StatsTracker
    from version_manager import VersionManager
    HAS_DEPENDENCIES = True
except ImportError:
    HAS_DEPENDENCIES = False


class HealthReporter:
    """Reporter for skills library health metrics."""

    def __init__(self):
        self.repo_root = Path.cwd()
        while not (self.repo_root / "package.json").exists() and self.repo_root != self.repo_root.parent:
            self.repo_root = self.repo_root.parent

        if HAS_DEPENDENCIES:
            self.version_manager = VersionManager()
            self.dep_checker = DependencyChecker()
            self.stats_tracker = StatsTracker()
        else:
            self.version_manager = None
            self.dep_checker = None
            self.stats_tracker = None

    def generate_health_report(self) -> Dict:
        """
        Generate comprehensive health report.

        Returns:
            Dictionary with health metrics
        """
        if not HAS_DEPENDENCIES:
            return {
                "error": "Required modules (dependency_checker, stats_tracker, version_manager) not available"
            }

        report = {
            "generated_at": datetime.now().isoformat(),
            "overview": {},
            "dependency_issues": [],
            "performance_issues": [],
            "recommendations": []
        }

        # Overview metrics
        skills = self.version_manager.list_skills()
        report['overview']['total_skills'] = len(skills)
        report['overview']['plugin_version'] = self.version_manager.get_plugin_version()

        # Check dependencies for all skills
        skills_with_deps = 0
        skills_with_missing_deps = 0

        for skill in skills:
            if 'dependencies' in skill and skill['dependencies']:
                skills_with_deps += 1

                dep_report = self.dep_checker.check_skill_dependencies(skill['name'])
                if dep_report.missing_python or dep_report.missing_mcp:
                    skills_with_missing_deps += 1

                    report['dependency_issues'].append({
                        "skill": skill['name'],
                        "missing_python": dep_report.missing_python,
                        "missing_mcp": dep_report.missing_mcp
                    })

        report['overview']['skills_with_dependencies'] = skills_with_deps
        report['overview']['skills_with_missing_dependencies'] = skills_with_missing_deps

        # Get performance statistics
        try:
            problematic_skills = self.stats_tracker.get_problematic_skills(10, 30)

            for stat in problematic_skills:
                if stat.success_rate < 80:
                    report['performance_issues'].append({
                        "skill": stat.skill_name,
                        "issue": "low_success_rate",
                        "success_rate": stat.success_rate,
                        "error_count": stat.error_count
                    })

            report['overview']['problematic_skills'] = len(problematic_skills)

        except Exception:
            report['overview']['problematic_skills'] = 0

        # Generate recommendations
        if skills_with_missing_deps > 0:
            report['recommendations'].append({
                "priority": "high",
                "category": "dependencies",
                "message": f"{skills_with_missing_deps} skills have missing dependencies. Run dependency checks and install missing packages."
            })

        if report['performance_issues']:
            report['recommendations'].append({
                "priority": "medium",
                "category": "performance",
                "message": f"{len(report['performance_issues'])} skills have performance issues. Review error logs and consider optimizations."
            })

        if not report['dependency_issues'] and not report['performance_issues']:
            report['recommendations'].append({
                "priority": "info",
                "category": "health",
                "message": "All skills are healthy! No issues detected."
            })

        return report

    def export_to_excel(self, output_path: str) -> bool:
        """
        Export health report to Excel file.

        Args:
            output_path: Path to output Excel file

        Returns:
            True if export succeeded
        """
        if not HAS_OPENPYXL:
            print("[ERROR] openpyxl library not installed. Install with: pip install openpyxl", file=sys.stderr)
            return False

        report = self.generate_health_report()

        if 'error' in report:
            print(f"[ERROR] {report['error']}", file=sys.stderr)
            return False

        # Create workbook
        wb = openpyxl.Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Create Overview sheet
        overview_sheet = wb.create_sheet("Overview")
        self._write_overview_sheet(overview_sheet, report['overview'])

        # Create Dependency Issues sheet
        if report['dependency_issues']:
            dep_sheet = wb.create_sheet("Dependency Issues")
            self._write_dependency_sheet(dep_sheet, report['dependency_issues'])

        # Create Performance Issues sheet
        if report['performance_issues']:
            perf_sheet = wb.create_sheet("Performance Issues")
            self._write_performance_sheet(perf_sheet, report['performance_issues'])

        # Create Recommendations sheet
        rec_sheet = wb.create_sheet("Recommendations")
        self._write_recommendations_sheet(rec_sheet, report['recommendations'])

        # Save workbook
        wb.save(output_path)
        print(f"[OK] Health report exported to: {output_path}")
        return True

    def _write_overview_sheet(self, sheet, overview: Dict):
        """Write overview data to sheet."""
        # Header
        sheet['A1'] = "Skills Library Health Report"
        sheet['A1'].font = Font(bold=True, size=14)

        sheet['A2'] = "Generated At"
        sheet['B2'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Metrics
        row = 4
        sheet[f'A{row}'] = "Metric"
        sheet[f'B{row}'] = "Value"
        sheet[f'A{row}'].font = Font(bold=True)
        sheet[f'B{row}'].font = Font(bold=True)

        row += 1
        for key, value in overview.items():
            sheet[f'A{row}'] = key.replace('_', ' ').title()
            sheet[f'B{row}'] = value
            row += 1

        # Auto-size columns
        sheet.column_dimensions['A'].width = 35
        sheet.column_dimensions['B'].width = 20

    def _write_dependency_sheet(self, sheet, issues: List[Dict]):
        """Write dependency issues to sheet."""
        # Header
        sheet['A1'] = "Dependency Issues"
        sheet['A1'].font = Font(bold=True, size=14)

        # Column headers
        headers = ['Skill', 'Missing Python Packages', 'Missing MCP Servers']
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Data
        row = 4
        for issue in issues:
            sheet.cell(row=row, column=1).value = issue['skill']
            sheet.cell(row=row, column=2).value = ', '.join(issue['missing_python']) if issue['missing_python'] else 'None'
            sheet.cell(row=row, column=3).value = ', '.join(issue['missing_mcp']) if issue['missing_mcp'] else 'None'
            row += 1

        # Auto-size
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 40
        sheet.column_dimensions['C'].width = 40

    def _write_performance_sheet(self, sheet, issues: List[Dict]):
        """Write performance issues to sheet."""
        # Header
        sheet['A1'] = "Performance Issues"
        sheet['A1'].font = Font(bold=True, size=14)

        # Column headers
        headers = ['Skill', 'Issue Type', 'Success Rate (%)', 'Error Count']
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Data
        row = 4
        for issue in issues:
            sheet.cell(row=row, column=1).value = issue['skill']
            sheet.cell(row=row, column=2).value = issue['issue'].replace('_', ' ').title()
            sheet.cell(row=row, column=3).value = f"{issue['success_rate']:.1f}"
            sheet.cell(row=row, column=4).value = issue['error_count']
            row += 1

        # Auto-size
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 25
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 15

    def _write_recommendations_sheet(self, sheet, recommendations: List[Dict]):
        """Write recommendations to sheet."""
        # Header
        sheet['A1'] = "Recommendations"
        sheet['A1'].font = Font(bold=True, size=14)

        # Column headers
        headers = ['Priority', 'Category', 'Recommendation']
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Data
        row = 4
        for rec in recommendations:
            sheet.cell(row=row, column=1).value = rec['priority'].upper()
            sheet.cell(row=row, column=2).value = rec['category'].title()
            sheet.cell(row=row, column=3).value = rec['message']
            sheet.cell(row=row, column=3).alignment = Alignment(wrap_text=True)
            row += 1

        # Auto-size
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 60


def main():
    """CLI entry point for testing."""
    if len(sys.argv) < 2:
        print("Usage: python health_reporter.py <report|excel> [output_path]")
        sys.exit(1)

    reporter = HealthReporter()
    command = sys.argv[1]

    if command == "report":
        report = reporter.generate_health_report()

        if 'error' in report:
            print(f"[ERROR] {report['error']}")
            sys.exit(1)

        print("\n=== Skills Health Report ===\n")
        print(f"Generated: {report['generated_at']}\n")

        print("Overview:")
        for key, value in report['overview'].items():
            print(f"  {key.replace('_', ' ').title()}: {value}")

        if report['dependency_issues']:
            print(f"\nDependency Issues ({len(report['dependency_issues'])}):")
            for issue in report['dependency_issues']:
                print(f"  - {issue['skill']}")
                if issue['missing_python']:
                    print(f"    Missing Python: {', '.join(issue['missing_python'])}")
                if issue['missing_mcp']:
                    print(f"    Missing MCP: {', '.join(issue['missing_mcp'])}")

        if report['performance_issues']:
            print(f"\nPerformance Issues ({len(report['performance_issues'])}):")
            for issue in report['performance_issues']:
                print(f"  - {issue['skill']}: {issue['issue']} (Success Rate: {issue['success_rate']:.1f}%)")

        print(f"\nRecommendations ({len(report['recommendations'])}):")
        for rec in report['recommendations']:
            print(f"  [{rec['priority'].upper()}] {rec['message']}")

    elif command == "excel":
        if len(sys.argv) < 3:
            print("Usage: python health_reporter.py excel <output_path>")
            sys.exit(1)

        output_path = sys.argv[2]
        success = reporter.export_to_excel(output_path)
        sys.exit(0 if success else 1)

    else:
        print(f"Unknown command: {command}")
        sys.exit(1)


if __name__ == "__main__":
    main()
